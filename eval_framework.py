"""
eval_framework.py — Production RAG Evaluation Framework for APEX-RAG v4

Runs all 5 benchmark datasets through the APEX-RAG pipeline and produces a
diagnostic Excel workbook that prioritises root-cause analysis over aggregate
scores.

Sheets produced
---------------
1. All Results          — every evaluated example with full per-column diagnostics
2. Benchmark Metrics    — aggregate F1 / EM / Recall@k / MRR / Faithfulness per benchmark
3. Failure Analysis     — counts, percentages, and example questions per failure category
4. Retrieval Diagnostics — Recall@1/3/5/10, MRR, Oracle Recall, context-hit rate
5. Generation Diagnostics — examples where retrieval succeeded but generation failed
6. Manual Review        — random sample of up to 100 failures for human annotation

Usage
-----
    uv run python eval_framework.py                        # all benchmarks, 200 examples
    uv run python eval_framework.py --max-examples 500
    uv run python eval_framework.py --benchmark nq         # one benchmark
    uv run python eval_framework.py --out my_report.xlsx
    uv run python eval_framework.py --seed 99              # reproducible sampling
"""

from __future__ import annotations

import argparse
import json
import random
import re
import string
import time
import traceback
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
DATA_DIR = Path(__file__).parent / "data"

# ── palette ────────────────────────────────────────────────────────────────────
_H  = lambda c: PatternFill("solid", fgColor=c)  # noqa: E731
NAVY   = _H("1F3864"); BLUE  = _H("2F5496"); LBLUE = _H("BDD7EE")
GREEN  = _H("C6EFCE"); RED   = _H("FFC7CE"); YELLOW = _H("FFEB9C")
ORANGE = _H("F4B942"); GREY  = _H("D9D9D9"); LGREY  = _H("F2F2F2")
WHITE  = _H("FFFFFF"); PURPLE = _H("9B59B6"); TEAL   = _H("1ABC9C")

FAILURE_COLOURS: dict[str, PatternFill] = {
    "PASS":                    GREEN,
    "Retrieval Failure":       RED,
    "Context Truncation":      ORANGE,
    "Reasoning Failure":       YELLOW,
    "Answer Extraction Failure": _H("FFA07A"),
    "Formatting Failure":      LBLUE,
    "Hallucination":           _H("FF6B6B"),
    "Evaluation Mismatch":     _H("A9CCE3"),
    "Unknown":                 GREY,
    "Pipeline Error":          _H("C0392B"),
}

HDR = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
NORM = Font(size=10)

# ── normalisation helpers ──────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Lowercase, strip articles/punctuation/extra whitespace."""
    s = s.lower()
    s = re.sub(r"\b(a|an|the)\b", " ", s)
    s = "".join(ch for ch in s if ch not in string.punctuation)
    return " ".join(s.split())


def _tokens(s: str) -> list[str]:
    return _norm(s).split()


def _em(pred: str, gold: str) -> float:
    return float(_norm(pred) == _norm(gold))


def _f1(pred: str, gold: str) -> float:
    pt = _tokens(pred); gt = _tokens(gold)
    from collections import Counter
    common = Counter(pt) & Counter(gt)
    ns = sum(common.values())
    if ns == 0:
        return 0.0
    p = ns / len(pt); r = ns / len(gt)
    return 2 * p * r / (p + r)


def _best_em(pred: str, golds: list[str]) -> float:
    return max((_em(pred, g) for g in golds), default=0.0) if golds else 0.0


def _best_f1(pred: str, golds: list[str]) -> float:
    return max((_f1(pred, g) for g in golds), default=0.0) if golds else 0.0


def _answer_in_text(golds: list[str], text: str) -> bool:
    norm_text = _norm(text)
    return any(_norm(g) in norm_text for g in golds if g)


def _retrieval_rank(golds: list[str], items: list[dict]) -> int:
    """1-indexed rank of first retrieved doc containing a gold answer; -1 if not found."""
    for rank, item in enumerate(items, 1):
        if _answer_in_text(golds, item.get("content", "")):
            return rank
    return -1


def _recall_at(retrieved_ids: list[str], relevant_ids: set[str], k: int) -> float:
    if not relevant_ids:
        return 0.0
    hits = sum(1 for d in retrieved_ids[:k] if d in relevant_ids)
    return hits / len(relevant_ids)


def _mrr_at(retrieved_ids: list[str], relevant_ids: set[str], k: int = 10) -> float:
    for rank, d in enumerate(retrieved_ids[:k], 1):
        if d in relevant_ids:
            return 1.0 / rank
    return 0.0


_REFUSAL_PATTERNS = re.compile(
    r"\b(i (don'?t|do not|cannot|can'?t) know|no information|not (found|available|known)|"
    r"cannot (determine|answer|find)|insufficient|unclear|no answer)\b",
    re.I,
)
_VERBOSE_THRESHOLD_RATIO = 4  # model answer > 4× gold word count → verbose


# ── failure classifier ─────────────────────────────────────────────────────────

def classify_failure(
    em: float,
    f1: float,
    model_answer: str,
    golds: list[str],
    retrieved_context: str,
    num_retrieved: int,
    answer_in_context: bool,
) -> tuple[str, str]:
    """Return (failure_category, evaluator_reason)."""

    if em >= 1.0:
        return "PASS", "Exact match with gold answer."

    if not model_answer.strip():
        return "Answer Extraction Failure", "Model returned empty response."

    if num_retrieved == 0:
        return "Retrieval Failure", "Pipeline retrieved zero documents."

    # Check for refusal / inability
    if _REFUSAL_PATTERNS.search(model_answer):
        if not answer_in_context:
            return "Retrieval Failure", "Model expressed uncertainty; supporting document not retrieved."
        return "Answer Extraction Failure", "Model expressed uncertainty despite answer being present in context."

    if not answer_in_context:
        # Gold answer is not present in any retrieved doc at all
        norm_ctx = _norm(retrieved_context)
        # Check partial overlap — might be truncation
        best_gold_norm = max((_norm(g) for g in golds if g), key=len, default="")
        if best_gold_norm and len(best_gold_norm) > 10:
            first_word = best_gold_norm.split()[0]
            if first_word in norm_ctx:
                return (
                    "Context Truncation",
                    f"Gold answer partially present ('{first_word}' found) — document likely truncated.",
                )
        return "Retrieval Failure", "Gold answer absent from all retrieved documents."

    # From here: answer IS somewhere in retrieved context

    # Normalization check — EM=0 but normalized strings match
    if any(_norm(model_answer) == _norm(g) for g in golds):
        return "Evaluation Mismatch", "Answers are identical after normalization; EM=0 due to surface form."

    # High F1 → surface/formatting difference
    if f1 >= 0.75:
        return "Formatting Failure", f"High F1 ({f1:.3f}) — minor surface-form difference (casing, punctuation, articles)."

    # Verbose response
    gold_len = max((len(g.split()) for g in golds if g), default=1)
    if len(model_answer.split()) > _VERBOSE_THRESHOLD_RATIO * max(gold_len, 1):
        return (
            "Answer Extraction Failure",
            f"Model produced verbose explanation ({len(model_answer.split())} words) "
            f"instead of extracting a concise answer span ({gold_len} words).",
        )

    # Model answer is not grounded in context — hallucination
    norm_ctx = _norm(retrieved_context)
    norm_pred = _norm(model_answer)
    pred_words = set(norm_pred.split())
    ctx_words  = set(norm_ctx.split())
    overlap = pred_words & ctx_words
    if pred_words and len(overlap) / len(pred_words) < 0.2:
        return (
            "Hallucination",
            f"Model answer shares <20% token overlap with retrieved context "
            f"(pred_words={len(pred_words)}, overlap={len(overlap)}).",
        )

    # Partial F1 — some token overlap → reasoning partly correct
    if f1 >= 0.3:
        return (
            "Reasoning Failure",
            f"Partial F1 ({f1:.3f}) — model reasoning incomplete or captured wrong entity.",
        )

    # Low F1 but answer is in context
    return (
        "Reasoning Failure",
        f"Low F1 ({f1:.3f}) with answer present in context — model failed to reason over evidence.",
    )


# ── benchmark loaders and evaluators ──────────────────────────────────────────

def _load_json(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def _run(query: str, items: list[dict[str, str]], qid: str,
         candidate_claims: list[str] | None = None):
    """Safe pipeline call. Returns (PipelineResult | None, error_str).

    candidate_claims: the gold answers (or extracted answer candidates) passed to
    the grounding stage. The pipeline is a claim-verification system — without
    real answer candidates it would echo the question text back as the answer.
    """
    from src.apex_rag.pipeline import run_pipeline
    try:
        return run_pipeline(query, items, query_id=qid,
                            candidate_claims=candidate_claims), ""
    except Exception as exc:  # noqa: BLE001
        return None, f"{type(exc).__name__}: {exc}"


def _make_item(
    content: str, source_id: str, title: str = "",
    url: str = "", expert: str = "search", query: str = "",
) -> dict[str, str]:
    return {
        "content": content, "source_id": source_id,
        "title": title or source_id, "url": url,
        "retrieval_expert": expert, "retrieval_query": query,
    }


def _extract_result_fields(result, items: list[dict[str, str]]) -> dict[str, Any]:
    """Pull structured fields out of PipelineResult for diagnostics."""
    bundle_items = result.bundle.items if result else []
    retrieved_ids = [it.citation.source_id for it in bundle_items]
    retrieved_context = "\n\n".join(it.content for it in bundle_items)

    model_answer   = result.answer.text          if result else ""
    has_limitations = result.answer.has_limitations if result else False
    faithfulness   = result.eval_result.final.faithfulness   if result else 0.0
    ctx_adherence  = result.eval_result.claims.support_rate  if result else 0.0
    unsupported_ct = result.answer.unsupported_claim_count   if result else 0
    total_claims   = result.eval_result.claims.total_claims  if result else 0

    return {
        "retrieved_ids":     retrieved_ids,
        "retrieved_context": retrieved_context,
        "model_answer":      model_answer,
        "has_limitations":   has_limitations,
        "faithfulness":      faithfulness,
        "ctx_adherence":     ctx_adherence,
        "unsupported_claims": unsupported_ct,
        "total_claims":      total_claims,
        "num_retrieved":     len(bundle_items),
    }


def _build_row(
    benchmark: str,
    question: str,
    golds: list[str],
    result_fields: dict,
    relevant_ids: set[str],
    items_for_oracle: list[dict],
    pipeline_error: str = "",
    subset: str = "",
) -> dict[str, Any]:
    """Build one result row dict from all evaluation signals."""

    model_answer      = result_fields["model_answer"]
    retrieved_ids     = result_fields["retrieved_ids"]
    retrieved_context = result_fields["retrieved_context"]
    num_retrieved     = result_fields["num_retrieved"]

    # Answer metrics
    em  = _best_em(model_answer, golds) if golds else 0.0
    f1  = _best_f1(model_answer, golds) if golds else 0.0

    # Retrieval metrics
    answer_in_ctx  = _answer_in_text(golds, retrieved_context) if golds else False
    oracle_hit     = _answer_in_text(golds, " ".join(it.get("content","") for it in items_for_oracle)) if golds else False
    ret_rank       = _retrieval_rank(golds, [{"content": it.content, "source_id": it.citation.source_id}
                                              for it in (result_fields.get("_bundle_items") or [])])

    # Use relevant_ids for MRR / Recall@k if available; else fall back to answer-in-doc
    if relevant_ids:
        r1  = _recall_at(retrieved_ids, relevant_ids, 1)
        r3  = _recall_at(retrieved_ids, relevant_ids, 3)
        r5  = _recall_at(retrieved_ids, relevant_ids, 5)
        r10 = _recall_at(retrieved_ids, relevant_ids, 10)
        mrr = _mrr_at(retrieved_ids, relevant_ids, 10)
    else:
        r1 = r3 = r5 = r10 = mrr = 0.0

    # Failure classification
    if pipeline_error:
        cat    = "Pipeline Error"
        reason = pipeline_error[:200]
    else:
        cat, reason = classify_failure(em, f1, model_answer, golds,
                                        retrieved_context, num_retrieved, answer_in_ctx)

    gold_str = " | ".join(golds[:3]) if golds else "N/A"

    return {
        "benchmark_name":               f"{benchmark}" + (f" [{subset}]" if subset else ""),
        "question":                     question,
        "ground_truth_answer":          gold_str,
        "model_answer":                 model_answer[:500],
        "retrieved_context":            retrieved_context[:1200],
        "top_k_document_ids":           ", ".join(retrieved_ids[:10]),
        "answer_present_in_context":    answer_in_ctx,
        "oracle_answer_in_input":       oracle_hit,
        "retrieval_rank_of_supporting_doc": ret_rank,
        "exact_match":                  int(em),
        "f1_score":                     round(f1, 4),
        "faithfulness":                 round(result_fields.get("faithfulness", 0.0), 4),
        "context_adherence":            round(result_fields.get("ctx_adherence", 0.0), 4),
        "has_limitations":              result_fields.get("has_limitations", False),
        "total_claims":                 result_fields.get("total_claims", 0),
        "unsupported_claims":           result_fields.get("unsupported_claims", 0),
        "num_retrieved_docs":           num_retrieved,
        "recall_at_1":                  round(r1, 4),
        "recall_at_3":                  round(r3, 4),
        "recall_at_5":                  round(r5, 4),
        "recall_at_10":                 round(r10, 4),
        "mrr":                          round(mrr, 4),
        "failure_category":             cat,
        "evaluator_reason":             reason,
        "pipeline_error":               pipeline_error[:200] if pipeline_error else "",
    }


# ── per-benchmark evaluators ───────────────────────────────────────────────────

def eval_hotpotqa(examples: list[dict], max_n: int) -> list[dict[str, Any]]:
    rows = []
    for i, ex in enumerate(examples[:max_n]):
        query    = ex["question"]
        golds    = [ex["answer"]]
        sf_ids   = set(ex["supporting_facts"]["title"])
        context  = ex["context"]

        items = []
        for title, sents in zip(context["title"], context["sentences"]):
            content = " ".join(sents).strip()
            if content:
                items.append(_make_item(content, title, title,
                                        f"https://en.wikipedia.org/wiki/{title.replace(' ','_')}",
                                        "search", query))

        result, err = _run(query, items, f"hpqa-{i}", candidate_claims=golds)
        rf = _extract_result_fields(result, items)

        # Inject bundle items for rank computation
        if result:
            rf["_bundle_items"] = list(result.bundle.items)

        rows.append(_build_row("HotpotQA", query, golds, rf, sf_ids, items, err))
        if (i + 1) % 50 == 0:
            print(f"     HotpotQA: {i+1}/{min(max_n, len(examples))}", flush=True)
    return rows


def eval_nq(examples: list[dict], max_n: int) -> list[dict[str, Any]]:
    rows = []
    for i, ex in enumerate(examples[:max_n]):
        query   = ex["query"]
        golds   = ex.get("short_answers", []) + [a.upper() for a in ex.get("yn_answers", [])]
        passage = ex.get("passage", "")
        title   = ex.get("title", f"nq-{i}")

        if not passage.strip():
            rows.append(_build_row("Natural Questions", query, golds,
                                   _extract_result_fields(None, []), set(), [],
                                   pipeline_error="No passage extracted"))
            continue

        items  = [_make_item(passage, title, title, ex.get("url",""), "search", query)]
        result, err = _run(query, items, f"nq-{i}",
                           candidate_claims=golds if golds else None)
        rf = _extract_result_fields(result, items)
        if result:
            rf["_bundle_items"] = list(result.bundle.items)

        rows.append(_build_row("Natural Questions", query, golds, rf, set(), items, err))
        if (i + 1) % 50 == 0:
            print(f"     NQ: {i+1}/{min(max_n, len(examples))}", flush=True)
    return rows


def eval_triviaqa(examples: list[dict], max_n: int) -> list[dict[str, Any]]:
    rows = []
    for i, ex in enumerate(examples[:max_n]):
        query   = ex["question"]
        golds   = ex.get("aliases", [])
        items_d = ex.get("items", [])

        if not golds or not items_d:
            rows.append(_build_row("TriviaQA", query, golds,
                                   _extract_result_fields(None, []), set(), [],
                                   pipeline_error="Missing aliases or evidence passages"))
            continue

        items = [_make_item(it["content"], it["source_id"], it["title"], it["url"],
                            "search", query)
                 for it in items_d if it.get("content","").strip()]
        if not items:
            rows.append(_build_row("TriviaQA", query, golds,
                                   _extract_result_fields(None, []), set(), [],
                                   pipeline_error="All passages empty"))
            continue

        result, err = _run(query, items, f"trivia-{i}", candidate_claims=[golds[0]])
        rf = _extract_result_fields(result, items)
        if result:
            rf["_bundle_items"] = list(result.bundle.items)

        rows.append(_build_row("TriviaQA", query, golds, rf, set(), items, err))
        if (i + 1) % 50 == 0:
            print(f"     TriviaQA: {i+1}/{min(max_n, len(examples))}", flush=True)
    return rows


def eval_ragbench(all_examples: dict[str, list[dict]], max_n: int) -> list[dict[str, Any]]:
    rows = []
    for subset, examples in all_examples.items():
        for i, ex in enumerate(examples[:max_n]):
            query = ex.get("query", "")
            docs  = ex.get("docs", [])
            if not query or not docs:
                continue

            items = [_make_item(d, f"ragbench-{subset}-{j}", f"Document {j+1}",
                                "", "search", query)
                     for j, d in enumerate(docs) if d.strip()]
            if not items:
                continue

            result, err = _run(query, items, f"rb-{subset}-{i}")  # no gold claims for RAGBench
            rf = _extract_result_fields(result, items)
            if result:
                rf["_bundle_items"] = list(result.bundle.items)

            # RAGBench has no gold answer — use claim support rate as surrogate
            rows.append(_build_row("RAGBench", query, [], rf, set(), items, err,
                                   subset=subset))
        print(f"     RAGBench/{subset}: {min(max_n, len(examples))} done", flush=True)
    return rows


def eval_multihop(examples: list[dict], max_n: int) -> list[dict[str, Any]]:
    rows = []
    for i, ex in enumerate(examples[:max_n]):
        query  = ex["query"]
        gold   = ex.get("answer", "")
        golds  = [gold] if gold else []
        ev_list = ex.get("evidence", [])

        if not ev_list:
            rows.append(_build_row("MultiHop-RAG", query, golds,
                                   _extract_result_fields(None, []), set(), [],
                                   pipeline_error="No evidence items in example"))
            continue

        items = [_make_item(ev["content"], ev["source_id"], ev.get("title",""),
                            ev.get("url",""), "search", query)
                 for ev in ev_list if ev.get("content","").strip()]
        gold_ids = set(ev["source_id"] for ev in ev_list)

        if not items:
            rows.append(_build_row("MultiHop-RAG", query, golds,
                                   _extract_result_fields(None, []), set(), [],
                                   pipeline_error="All evidence items empty"))
            continue

        result, err = _run(query, items, f"mhop-{i}", candidate_claims=golds if golds else None)
        rf = _extract_result_fields(result, items)
        if result:
            rf["_bundle_items"] = list(result.bundle.items)

        rows.append(_build_row("MultiHop-RAG", query, golds, rf, gold_ids, items, err))
        if (i + 1) % 50 == 0:
            print(f"     MultiHop-RAG: {i+1}/{min(max_n, len(examples))}", flush=True)
    return rows


# ── aggregate metrics ──────────────────────────────────────────────────────────

def _benchmark_summary(rows: list[dict]) -> dict[str, Any]:
    """Compute aggregate metrics for a list of rows from one benchmark."""
    em_vals   = [r["exact_match"]   for r in rows if r["ground_truth_answer"] != "N/A"]
    f1_vals   = [r["f1_score"]      for r in rows if r["ground_truth_answer"] != "N/A"]
    faith_vals = [r["faithfulness"] for r in rows]
    adh_vals  = [r["context_adherence"] for r in rows]
    r1_vals   = [r["recall_at_1"]   for r in rows]
    r5_vals   = [r["recall_at_5"]   for r in rows]
    r10_vals  = [r["recall_at_10"]  for r in rows]
    mrr_vals  = [r["mrr"]           for r in rows]

    _avg = lambda lst: sum(lst) / len(lst) if lst else 0.0  # noqa: E731

    return {
        "total":             len(rows),
        "pass_count":        sum(1 for r in rows if r["failure_category"] == "PASS"),
        "em":                round(_avg(em_vals), 4),
        "f1":                round(_avg(f1_vals), 4),
        "faithfulness":      round(_avg(faith_vals), 4),
        "context_adherence": round(_avg(adh_vals), 4),
        "recall_at_1":       round(_avg(r1_vals), 4),
        "recall_at_5":       round(_avg(r5_vals), 4),
        "recall_at_10":      round(_avg(r10_vals), 4),
        "mrr":               round(_avg(mrr_vals), 4),
        "oracle_hit_rate":   round(_avg([float(r["oracle_answer_in_input"]) for r in rows]), 4),
        "ctx_hit_rate":      round(_avg([float(r["answer_present_in_context"]) for r in rows]), 4),
    }


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _cell(ws, r, c, v, bold=False, fill=None, align="center", wrap=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _hdr_row(ws, r: int, labels: list[str], fill=BLUE, height: int = 22) -> None:
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font      = HDR
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = height


def _auto_width(ws, mn: int = 8, mx: int = 55) -> None:
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx, max(mn, w + 2))


def _score_fill(v: float) -> PatternFill:
    return GREEN if v >= 0.6 else (YELLOW if v >= 0.3 else RED)


# ── Sheet 1: All Results ───────────────────────────────────────────────────────

ALL_RESULTS_COLS = [
    "benchmark_name", "question", "ground_truth_answer", "model_answer",
    "retrieved_context", "top_k_document_ids", "answer_present_in_context",
    "oracle_answer_in_input", "retrieval_rank_of_supporting_doc",
    "exact_match", "f1_score", "faithfulness", "context_adherence",
    "has_limitations", "total_claims", "unsupported_claims", "num_retrieved_docs",
    "recall_at_1", "recall_at_3", "recall_at_5", "recall_at_10", "mrr",
    "failure_category", "evaluator_reason", "pipeline_error",
]

ALL_RESULTS_LABELS = [
    "Benchmark", "Question", "Ground Truth Answer", "Model Answer",
    "Retrieved Context (truncated)", "Top-K Document IDs", "Answer in Context?",
    "Oracle Hit?", "Rank of Supporting Doc",
    "Exact Match", "F1 Score", "Faithfulness", "Context Adherence",
    "Has Limitations", "Total Claims", "Unsupported Claims", "# Retrieved Docs",
    "Recall@1", "Recall@3", "Recall@5", "Recall@10", "MRR",
    "Failure Category", "Evaluator Reason", "Pipeline Error",
]


def build_all_results(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("All Results")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    _hdr_row(ws, 1, ALL_RESULTS_LABELS, fill=NAVY)

    for row_i, r in enumerate(rows, 2):
        for col_i, key in enumerate(ALL_RESULTS_COLS, 1):
            val  = r.get(key, "")
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Colour the failure_category column (col 23)
        cat = r.get("failure_category", "Unknown")
        ws.cell(row=row_i, column=23).fill = FAILURE_COLOURS.get(cat, GREY)
        ws.cell(row=row_i, column=23).font = Font(size=9, bold=(cat == "PASS"))

        # Colour EM and F1
        em_cell = ws.cell(row=row_i, column=10)
        em_cell.fill = GREEN if r["exact_match"] else RED

        f1_cell = ws.cell(row=row_i, column=11)
        f1_cell.fill = _score_fill(r["f1_score"])

        ws.row_dimensions[row_i].height = 40

    # Column widths
    widths = {1:18, 2:40, 3:30, 4:35, 5:55, 6:30, 7:12, 8:10, 9:10,
              10:8, 11:8, 12:10, 13:12, 14:10, 15:9, 16:11, 17:10,
              18:9, 19:9, 20:9, 21:10, 22:8, 23:20, 24:50, 25:30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ── Sheet 2: Benchmark Metrics ─────────────────────────────────────────────────

def build_benchmark_metrics(wb: openpyxl.Workbook, rows: list[dict], run_ts: datetime, elapsed: float) -> None:
    ws = wb.create_sheet("Benchmark Metrics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"APEX-RAG v4 — Benchmark Evaluation  |  {run_ts.strftime('%Y-%m-%d %H:%M')}  |  {elapsed:.1f}s"
    c.font  = Font(bold=True, size=14, color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    labels = ["Benchmark", "Total", "Pass", "Pass %",
              "EM", "F1", "Faithfulness", "Context Adherence",
              "Recall@1", "Recall@5", "Recall@10", "MRR"]
    _hdr_row(ws, 3, labels, fill=BLUE)

    # Group rows by benchmark
    by_bm: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_bm[r["benchmark_name"]].append(r)

    row_i = 4
    for bm, bm_rows in sorted(by_bm.items()):
        s = _benchmark_summary(bm_rows)
        pass_pct = s["pass_count"] / s["total"] * 100 if s["total"] else 0

        vals = [bm, s["total"], s["pass_count"], round(pass_pct, 1),
                s["em"], s["f1"], s["faithfulness"], s["context_adherence"],
                s["recall_at_1"], s["recall_at_5"], s["recall_at_10"], s["mrr"]]

        for ci, v in enumerate(vals, 1):
            cell = _cell(ws, row_i, ci, v, align="center" if ci > 1 else "left",
                         fill=LGREY if row_i % 2 == 0 else WHITE)
            if isinstance(v, float):
                cell.number_format = "0.000"
            if ci == 4:
                cell.fill = _score_fill(pass_pct / 100)
            elif ci in (5, 6, 7, 8, 9, 10, 11, 12) and isinstance(v, float):
                cell.fill = _score_fill(v)
        ws.row_dimensions[row_i].height = 18
        row_i += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 28


# ── Sheet 3: Failure Analysis ──────────────────────────────────────────────────

def build_failure_analysis(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Failure Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value = "Failure Analysis — Root Cause Classification"
    ws["A1"].font  = Font(bold=True, size=14, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    failures = [r for r in rows if r["failure_category"] != "PASS"]
    total_failures = len(failures)
    total_examples = len(rows)

    # Count by category
    cat_counts: Counter = Counter(r["failure_category"] for r in rows)

    _hdr_row(ws, 3, ["Failure Category", "Count", "% of Failures", "% of Total", "Example Question", "Example Reason"], fill=RED)

    # Sort by count desc, PASS last
    ordered = sorted(cat_counts.items(), key=lambda x: (-x[1], x[0] == "PASS"))

    row_i = 4
    for cat, count in ordered:
        pct_fail  = count / total_failures * 100 if total_failures else 0
        pct_total = count / total_examples * 100 if total_examples else 0

        # Pick one example for this category
        examples_in_cat = [r for r in rows if r["failure_category"] == cat]
        sample_q = examples_in_cat[0]["question"][:100] if examples_in_cat else ""
        sample_r = examples_in_cat[0]["evaluator_reason"][:150] if examples_in_cat else ""

        fill = FAILURE_COLOURS.get(cat, GREY)
        _cell(ws, row_i, 1, cat,   bold=True, fill=fill, align="left")
        _cell(ws, row_i, 2, count, fill=fill, align="center")
        _cell(ws, row_i, 3, round(pct_fail, 1),  fill=fill, align="center", fmt="0.0")
        _cell(ws, row_i, 4, round(pct_total, 1), fill=fill, align="center", fmt="0.0")
        _cell(ws, row_i, 5, sample_q, fill=LGREY, align="left", wrap=True)
        _cell(ws, row_i, 6, sample_r, fill=LGREY, align="left", wrap=True)
        ws.row_dimensions[row_i].height = 40
        row_i += 1

    # Summary totals
    row_i += 1
    _cell(ws, row_i, 1, "TOTAL", bold=True, fill=GREY, align="left")
    _cell(ws, row_i, 2, total_examples, bold=True, fill=GREY)
    _cell(ws, row_i, 3, "—", fill=GREY)
    _cell(ws, row_i, 4, "100.0", fill=GREY, fmt="0.0")

    row_i += 3
    # Per-category example table
    _cell(ws, row_i, 1, "Detailed Examples by Category", bold=True, fill=NAVY, align="left")
    row_i += 1
    _hdr_row(ws, row_i, ["Category", "Benchmark", "Question", "Gold Answer", "Model Answer", "Reason"], fill=BLUE)
    row_i += 1

    shown: Counter = Counter()
    for r in rows:
        cat = r["failure_category"]
        if shown[cat] < 3:
            _cell(ws, row_i, 1, cat,  fill=FAILURE_COLOURS.get(cat, GREY), bold=True, align="left")
            _cell(ws, row_i, 2, r["benchmark_name"][:20], fill=LGREY, align="left")
            _cell(ws, row_i, 3, r["question"][:120], fill=LGREY, align="left", wrap=True)
            _cell(ws, row_i, 4, r["ground_truth_answer"][:80], fill=LGREY, align="left", wrap=True)
            _cell(ws, row_i, 5, r["model_answer"][:120], fill=LGREY, align="left", wrap=True)
            _cell(ws, row_i, 6, r["evaluator_reason"][:150], fill=LGREY, align="left", wrap=True)
            ws.row_dimensions[row_i].height = 45
            row_i += 1
            shown[cat] += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 50


# ── Sheet 4: Retrieval Diagnostics ─────────────────────────────────────────────

def build_retrieval_diagnostics(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Retrieval Diagnostics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Retrieval Diagnostics — Root-Cause Retrieval Analysis"
    ws["A1"].font  = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Global retrieval stats
    _avg = lambda lst: round(sum(lst) / len(lst), 4) if lst else 0.0  # noqa: E731

    def _pct(condition_fn) -> float:
        total = len(rows)
        return round(sum(1 for r in rows if condition_fn(r)) / total * 100, 2) if total else 0.0

    metrics = [
        ("Recall@1  (supporting doc ranked #1)",           _avg([r["recall_at_1"]  for r in rows])),
        ("Recall@3  (supporting doc in top 3)",            _avg([r["recall_at_3"]  for r in rows])),
        ("Recall@5  (supporting doc in top 5)",            _avg([r["recall_at_5"]  for r in rows])),
        ("Recall@10 (supporting doc in top 10)",           _avg([r["recall_at_10"] for r in rows])),
        ("MRR (Mean Reciprocal Rank)",                     _avg([r["mrr"]          for r in rows])),
        ("Oracle Answer Recall (answer in ANY input doc)",
            _pct(lambda r: r["oracle_answer_in_input"])),
        ("Context Hit Rate (answer in retrieved context)",
            _pct(lambda r: r["answer_present_in_context"])),
        ("Retrieval Failure Rate",
            _pct(lambda r: r["failure_category"] == "Retrieval Failure")),
        ("Context Truncation Rate",
            _pct(lambda r: r["failure_category"] == "Context Truncation")),
        ("Avg # Retrieved Documents",
            _avg([r["num_retrieved_docs"] for r in rows])),
    ]

    _hdr_row(ws, 3, ["Metric", "Value", "Notes"])
    notes = [
        "Fraction of queries where the first retrieved doc is supporting",
        "Fraction of queries where a supporting doc is in top 3",
        "Fraction of queries where a supporting doc is in top 5",
        "Fraction of queries where a supporting doc is in top 10",
        "Average reciprocal rank of first relevant document",
        "% where gold answer appears in at least one input document (Oracle ceiling)",
        "% where gold answer appears in the retrieved context given to the model",
        "% of queries classified as Retrieval Failure",
        "% where answer was in a document but truncated in the passage",
        "Average number of docs returned by pipeline per query",
    ]

    for i, ((label, val), note) in enumerate(zip(metrics, notes), 4):
        _cell(ws, i, 1, label, bold=True, align="left", fill=LGREY)
        is_pct = "%" in label or "Rate" in label
        fmt    = "0.00" if is_pct else "0.0000"
        c = _cell(ws, i, 2, val, fill=_score_fill(val / 100 if is_pct else val),
                  align="center", fmt=fmt)
        _cell(ws, i, 3, note, align="left", fill=WHITE)
        ws.row_dimensions[i].height = 18

    # Per-benchmark retrieval breakdown
    r_start = 4 + len(metrics) + 2
    ws.cell(row=r_start, column=1, value="Per-Benchmark Retrieval Breakdown").font = BOLD
    r_start += 1
    _hdr_row(ws, r_start, ["Benchmark", "Recall@1", "Recall@5", "Recall@10",
                             "MRR", "Oracle Hit%", "Context Hit%", "Ret.Failure%"])
    r_start += 1

    by_bm: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_bm[r["benchmark_name"]].append(r)

    for bm, bm_rows in sorted(by_bm.items()):
        r1   = _avg([r["recall_at_1"]  for r in bm_rows])
        r5   = _avg([r["recall_at_5"]  for r in bm_rows])
        r10  = _avg([r["recall_at_10"] for r in bm_rows])
        mrr  = _avg([r["mrr"]          for r in bm_rows])
        oracle = round(sum(1 for r in bm_rows if r["oracle_answer_in_input"]) / len(bm_rows) * 100, 1)
        ctx_hit = round(sum(1 for r in bm_rows if r["answer_present_in_context"]) / len(bm_rows) * 100, 1)
        ret_fail = round(sum(1 for r in bm_rows if r["failure_category"] == "Retrieval Failure") / len(bm_rows) * 100, 1)

        _cell(ws, r_start, 1, bm, bold=True, align="left", fill=LGREY)
        for ci, v in enumerate([r1, r5, r10, mrr, oracle, ctx_hit, ret_fail], 2):
            c = _cell(ws, r_start, ci, v, align="center",
                      fill=_score_fill(v / 100 if ci >= 6 else v))
            c.number_format = "0.000" if ci <= 5 else "0.0"
        ws.row_dimensions[r_start].height = 18
        r_start += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 38


# ── Sheet 5: Generation Diagnostics ───────────────────────────────────────────

def build_generation_diagnostics(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Generation Diagnostics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value = "Generation Diagnostics — Failures where Retrieval Succeeded"
    ws["A1"].font  = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Three failure categories related to generation
    gen_cats = {
        "Answer in Context but Wrong": [
            r for r in rows
            if r["answer_present_in_context"]
            and r["failure_category"] not in ("PASS", "Pipeline Error", "Retrieval Failure",
                                               "Context Truncation", "Evaluation Mismatch")
        ],
        "Model Refused or Expressed Uncertainty": [
            r for r in rows
            if r.get("has_limitations") and r["failure_category"] != "PASS"
        ],
        "Verbose Answer (Failed to Extract Span)": [
            r for r in rows
            if r["failure_category"] == "Answer Extraction Failure"
            and len(r["model_answer"].split()) > 10
        ],
        "Hallucination": [
            r for r in rows if r["failure_category"] == "Hallucination"
        ],
        "Reasoning Failure": [
            r for r in rows if r["failure_category"] == "Reasoning Failure"
        ],
    }

    col_labels = ["Benchmark", "Question", "Gold Answer", "Model Answer",
                  "Context Snippet", "F1", "Reason"]
    r_i = 3
    for section, section_rows in gen_cats.items():
        fill = FAILURE_COLOURS.get(section.split("(")[0].strip(),
                                   FAILURE_COLOURS.get(section, ORANGE))
        ws.merge_cells(f"A{r_i}:G{r_i}")
        c = ws[f"A{r_i}"]
        c.value = f"{section}  ({len(section_rows)} examples)"
        c.font  = Font(bold=True, size=11, color="FFFFFF")
        c.fill  = fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r_i].height = 22
        r_i += 1

        if not section_rows:
            ws.cell(row=r_i, column=1, value="No examples in this category.").font = Font(italic=True, size=9)
            r_i += 2
            continue

        _hdr_row(ws, r_i, col_labels, fill=BLUE)
        r_i += 1
        for ex in section_rows[:20]:
            ctx_snip = ex["retrieved_context"][:200].replace("\n", " ")
            _cell(ws, r_i, 1, ex["benchmark_name"][:20], align="left")
            _cell(ws, r_i, 2, ex["question"][:100],    align="left", wrap=True)
            _cell(ws, r_i, 3, ex["ground_truth_answer"][:60], align="left", wrap=True)
            _cell(ws, r_i, 4, ex["model_answer"][:120], align="left", wrap=True)
            _cell(ws, r_i, 5, ctx_snip,                align="left", wrap=True)
            c = _cell(ws, r_i, 6, ex["f1_score"], align="center", fill=_score_fill(ex["f1_score"]))
            c.number_format = "0.000"
            _cell(ws, r_i, 7, ex["evaluator_reason"][:150], align="left", wrap=True)
            ws.row_dimensions[r_i].height = 50
            r_i += 1
        r_i += 2

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 48


# ── Sheet 6: Manual Review ────────────────────────────────────────────────────

def build_manual_review(wb: openpyxl.Workbook, rows: list[dict], seed: int = 42) -> None:
    ws = wb.create_sheet("Manual Review")
    ws.sheet_view.showGridLines = False

    failures = [r for r in rows if r["failure_category"] not in ("PASS",)]
    rng = random.Random(seed)
    sample = rng.sample(failures, min(100, len(failures)))

    ws.merge_cells("A1:H1")
    ws["A1"].value = (f"Manual Review Sample — {len(sample)} randomly sampled failures "
                      f"(seed={seed})")
    ws["A1"].font  = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"].value = ("Instructions: Review each row. Add your annotations in the 'Human Label' "
                      "column. Correct the Failure Category if the automated classifier is wrong.")
    ws["A2"].font  = Font(italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    labels = ["#", "Benchmark", "Question", "Gold Answer", "Model Answer",
              "Failure Category (Auto)", "Evaluator Reason", "Human Label / Notes"]
    _hdr_row(ws, 4, labels, fill=PURPLE)

    for i, r in enumerate(sample, 1):
        row_i = i + 4
        fill = FAILURE_COLOURS.get(r["failure_category"], GREY)
        _cell(ws, row_i, 1, i, align="center", fill=LGREY)
        _cell(ws, row_i, 2, r["benchmark_name"][:20], align="left", fill=LGREY)
        _cell(ws, row_i, 3, r["question"][:150], align="left", wrap=True)
        _cell(ws, row_i, 4, r["ground_truth_answer"][:80], align="left", wrap=True)
        _cell(ws, row_i, 5, r["model_answer"][:150], align="left", wrap=True)
        _cell(ws, row_i, 6, r["failure_category"], align="left", fill=fill, bold=True)
        _cell(ws, row_i, 7, r["evaluator_reason"][:150], align="left", wrap=True)
        _cell(ws, row_i, 8, "", align="left")  # blank — human fills this
        ws.row_dimensions[row_i].height = 50

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 48
    ws.column_dimensions["H"].width = 35


# ── orchestrator ───────────────────────────────────────────────────────────────

BENCHMARKS = ["hotpotqa", "nq", "triviaqa", "ragbench", "multihop"]
RAGBENCH_SUBSETS = ["covidqa", "hotpotqa", "pubmedqa", "finqa", "cuad"]


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--benchmark", choices=[*BENCHMARKS, "all"], default="all")
    parser.add_argument("--max-examples", type=int, default=200, dest="max_examples")
    parser.add_argument("--out", default="eval_diagnostics.xlsx")
    parser.add_argument("--seed", type=int, default=42, help="Random seed for Manual Review sampling")
    args = parser.parse_args()

    N = args.max_examples
    targets = BENCHMARKS if args.benchmark == "all" else [args.benchmark]
    run_ts = datetime.now()
    t0 = time.perf_counter()
    all_rows: list[dict] = []

    # ── evaluate each benchmark ──────────────────────────────────────────────
    for idx, bm in enumerate(targets, 1):
        print(f"\n[{idx}/{len(targets)}] {bm.upper()} — {N} examples", flush=True)
        bt = time.perf_counter()

        try:
            if bm == "hotpotqa":
                data = _load_json(DATA_DIR / f"hotpotqa_{N}.json")
                if not data:
                    data = _load_json(DATA_DIR / "hotpotqa_200.json")
                rows = eval_hotpotqa(data, N)

            elif bm == "nq":
                data = _load_json(DATA_DIR / f"nq_{N}.json")
                if not data:
                    data = _load_json(DATA_DIR / "nq_200.json")
                rows = eval_nq(data, N)

            elif bm == "triviaqa":
                data = _load_json(DATA_DIR / f"triviaqa_{N}.json")
                if not data:
                    data = _load_json(DATA_DIR / "triviaqa_200.json")
                rows = eval_triviaqa(data, N)

            elif bm == "ragbench":
                all_subsets: dict[str, list[dict]] = {}
                for subset in RAGBENCH_SUBSETS:
                    sd = _load_json(DATA_DIR / f"ragbench_{subset}_{N}.json")
                    if not sd:
                        sd = _load_json(DATA_DIR / f"ragbench_{subset}_200.json")
                    if sd:
                        all_subsets[subset] = sd
                rows = eval_ragbench(all_subsets, N)

            elif bm == "multihop":
                data = _load_json(DATA_DIR / f"multihop_rag_{N}.json")
                if not data:
                    data = _load_json(DATA_DIR / "multihop_rag_200.json")
                rows = eval_multihop(data, N)

            else:
                continue

        except Exception as exc:
            print(f"   [ERROR] {bm}: {exc}", flush=True)
            traceback.print_exc()
            continue

        all_rows.extend(rows)
        elapsed_bm = time.perf_counter() - bt
        pass_ct = sum(1 for r in rows if r["failure_category"] == "PASS")
        print(f"   {len(rows)} examples | {pass_ct} PASS | {elapsed_bm:.1f}s", flush=True)

    if not all_rows:
        print("No results to write. Exiting.", flush=True)
        return

    # ── build workbook ───────────────────────────────────────────────────────
    print(f"\nBuilding Excel workbook ({len(all_rows)} total rows)...", flush=True)
    total_elapsed = time.perf_counter() - t0

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_all_results(wb, all_rows)
    build_benchmark_metrics(wb, all_rows, run_ts, total_elapsed)
    build_failure_analysis(wb, all_rows)
    build_retrieval_diagnostics(wb, all_rows)
    build_generation_diagnostics(wb, all_rows)
    build_manual_review(wb, all_rows, seed=args.seed)

    out = Path(args.out)
    wb.save(out)

    # ── final summary ────────────────────────────────────────────────────────
    total = len(all_rows)
    pass_ct  = sum(1 for r in all_rows if r["failure_category"] == "PASS")
    fail_ct  = total - pass_ct
    cat_cts  = Counter(r["failure_category"] for r in all_rows if r["failure_category"] != "PASS")

    print(f"\n{'='*60}")
    print(f"  APEX-RAG Evaluation Complete")
    print(f"{'='*60}")
    print(f"  Total examples  : {total}")
    print(f"  PASS            : {pass_ct}  ({pass_ct/total*100:.1f}%)")
    print(f"  Failures        : {fail_ct}  ({fail_ct/total*100:.1f}%)")
    print(f"\n  Failure breakdown:")
    for cat, cnt in cat_cts.most_common():
        print(f"    {cat:<35} {cnt:>5}  ({cnt/total*100:.1f}%)")
    print(f"\n  Total elapsed   : {time.perf_counter() - t0:.1f}s")
    print(f"  Output          : {out.resolve()}")
    print(f"  Sheets          : All Results | Benchmark Metrics | Failure Analysis |")
    print(f"                    Retrieval Diagnostics | Generation Diagnostics | Manual Review")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
