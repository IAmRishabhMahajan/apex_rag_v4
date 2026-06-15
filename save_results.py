"""Parse test_results.json and write a structured Excel workbook."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── colours ──────────────────────────────────────────────────────────────────
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="BDD7EE")
GREY = PatternFill("solid", fgColor="D9D9D9")
WHITE = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FONT = Font(bold=True, size=10)

# ── helpers ───────────────────────────────────────────────────────────────────


def _h(ws, row: int, col: int, value, bold: bool = False, fill=None, align: str = "center") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bold:
        cell.font = Font(bold=True, size=10)
    if fill:
        cell.fill = fill


def _header_row(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 12, max_w: int = 48) -> None:
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max_w, max(min_w, length + 2)
        )


# ── module groupings ──────────────────────────────────────────────────────────
_US_MAP = {
    "US-001": "Query Intelligence",
    "US-002": "Adaptive Retrieval Planning",
    "US-003": "Expert Retrieval Routing",
    "US-004": "Evidence Fusion",
    "US-005": "Validation Mesh",
    "US-006": "Complex Query Reasoning",
    "US-007": "Evidence Scoring & Gap Detection",
    "US-008": "Retrieval Repair Loop",
    "US-009": "Grounded Generation",
    "US-010": "Risk Verification",
    "US-011": "APEX-Eval Framework",
    "US-012": "Research Traceability",
    "US-013": "HotpotQA Evaluation",
    "US-014": "MultiHop-RAG Evaluation",
    "US-015": "RAGBench Evaluation",
    "US-016": "Open RAG Bench Evaluation",
    "US-017": "EnterpriseRAG-Bench",
    "US-018": "Natural Questions",
    "US-019": "TriviaQA",
    "US-020": "FanOutQA",
    "US-021": "T²-RAGBench",
    "US-022": "MEBench",
    "US-023": "MuDABench",
}

_MODULE_MAP = {
    "test_us_001": "US-001",
    "test_us_002": "US-002",
    "test_us_003": "US-003",
    "test_us_004": "US-004",
    "test_us_005": "US-005",
    "test_us_006": "US-006",
    "test_us_007": "US-007",
    "test_us_008": "US-008",
    "test_us_009": "US-009",
    "test_us_010": "US-010",
    "test_us_011": "US-011",
    "test_us_012": "US-012",
    "test_pipeline": "pipeline",
    "story_checks": "story_checks",
    "test_story_index": "story_index",
}


def _us_from_nodeid(nodeid: str) -> str:
    """Extract US-XXX tag from a test node ID."""
    for prefix, us in _MODULE_MAP.items():
        if prefix in nodeid:
            return us
    return "other"


def _status_fill(outcome: str):
    return GREEN if outcome == "passed" else RED if outcome == "failed" else YELLOW


# ═══════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════════════════════


def _sheet_summary(wb: openpyxl.Workbook, data: dict) -> None:
    """Overview sheet: headline numbers + run metadata."""
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    total = data["summary"]["total"]
    passed = data["summary"]["passed"]
    failed = data["summary"].get("failed", 0)
    errors = data["summary"].get("error", 0)
    skipped = data["summary"].get("skipped", 0)
    duration = data.get("duration", 0.0)
    run_ts = datetime.fromtimestamp(data.get("created", datetime.now().timestamp()))

    # Title
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "APEX-RAG v4 — Test Results"
    cell.font = Font(bold=True, size=16, color="2F5496")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Metadata row
    ws.merge_cells("A2:F2")
    meta = ws["A2"]
    meta.value = f"Run: {run_ts.strftime('%Y-%m-%d %H:%M')}   |   Duration: {duration:.2f}s"
    meta.font = Font(italic=True, size=10, color="595959")
    meta.alignment = Alignment(horizontal="center", vertical="center")

    # Scorecards
    cards = [
        ("Total", total, BLUE),
        ("Passed", passed, GREEN),
        ("Failed", failed, RED if failed else WHITE),
        ("Errors", errors, RED if errors else WHITE),
        ("Skipped", skipped, YELLOW if skipped else WHITE),
        (
            "Pass %",
            f"{passed / total * 100:.1f}%" if total else "—",
            GREEN if failed == 0 else YELLOW,
        ),
    ]
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 36
    for col, (label, value, fill) in enumerate(cards, 1):
        _h(ws, 4, col, label, bold=True, fill=GREY)
        cell = ws.cell(row=5, column=col, value=value)
        cell.fill = fill
        cell.font = Font(bold=True, size=18)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Per-module breakdown header
    ws.row_dimensions[7].height = 20
    _header_row(ws, 7, ["Module / Story", "Story Name", "Total", "Passed", "Failed", "Pass %"])

    # Group tests by module
    by_module: dict[str, dict] = {}
    for test in data["tests"]:
        us = _us_from_nodeid(test["nodeid"])
        if us not in by_module:
            by_module[us] = {"total": 0, "passed": 0, "failed": 0}
        by_module[us]["total"] += 1
        if test["outcome"] == "passed":
            by_module[us]["passed"] += 1
        elif test["outcome"] in ("failed", "error"):
            by_module[us]["failed"] += 1

    row = 8
    for us_key in sorted(by_module.keys()):
        m = by_module[us_key]
        pct = m["passed"] / m["total"] * 100 if m["total"] else 0
        fill = GREEN if m["failed"] == 0 else RED
        story_name = _US_MAP.get(us_key, "—")
        ws.cell(row=row, column=1, value=us_key).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=story_name).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=3, value=m["total"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=m["passed"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=m["failed"]).alignment = Alignment(horizontal="center")
        pct_cell = ws.cell(row=row, column=6, value=f"{pct:.0f}%")
        pct_cell.fill = fill
        pct_cell.alignment = Alignment(horizontal="center")
        if m["failed"] > 0:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = RED
        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34


def _sheet_all_tests(wb: openpyxl.Workbook, data: dict) -> None:
    """Full test list with outcome, duration, and node ID."""
    ws = wb.create_sheet("All Tests")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    _header_row(ws, 1, ["#", "Story", "Test Class", "Test Name", "Outcome", "Duration (s)"])

    for i, test in enumerate(data["tests"], 1):
        nodeid: str = test["nodeid"]
        outcome: str = test["outcome"]
        duration: float = test.get("duration", 0.0)
        us = _us_from_nodeid(nodeid)

        # Parse class + test name from nodeid
        parts = nodeid.split("::")
        test_class = parts[1] if len(parts) >= 3 else "—"
        test_name = parts[-1]

        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=us).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=test_class).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=4, value=test_name).alignment = Alignment(horizontal="left")
        fill = _status_fill(outcome)
        cell = ws.cell(row=row, column=5, value=outcome.upper())
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=round(duration, 4)).alignment = Alignment(
            horizontal="right"
        )

    _auto_width(ws)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12


def _sheet_failures(wb: openpyxl.Workbook, data: dict) -> None:
    """Dedicated sheet for any failed tests with error details."""
    ws = wb.create_sheet("Failures")
    ws.sheet_view.showGridLines = False

    failed_tests = [t for t in data["tests"] if t["outcome"] in ("failed", "error")]

    if not failed_tests:
        ws.merge_cells("A1:C1")
        cell = ws["A1"]
        cell.value = "✓ All tests passed — no failures to show."
        cell.font = Font(bold=True, color="375623", size=12)
        cell.fill = GREEN
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        return

    _header_row(ws, 1, ["Story", "Test Node ID", "Outcome", "Error Message"])
    row = 2
    for test in failed_tests:
        nodeid = test["nodeid"]
        outcome = test["outcome"]
        us = _us_from_nodeid(nodeid)

        # Extract short error message
        call = test.get("call") or {}
        longrepr = call.get("longrepr", "") or ""
        # Take last non-empty line as the short error
        lines = [ln.strip() for ln in longrepr.splitlines() if ln.strip()]
        short_err = lines[-1] if lines else "—"

        ws.cell(row=row, column=1, value=us).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=nodeid).alignment = Alignment(horizontal="left")
        cell = ws.cell(row=row, column=3, value=outcome.upper())
        cell.fill = RED
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=short_err).alignment = Alignment(
            horizontal="left", wrap_text=True
        )
        ws.row_dimensions[row].height = 30
        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 70


def _sheet_by_story(wb: openpyxl.Workbook, data: dict) -> None:
    """One section per US story with all its test results."""
    ws = wb.create_sheet("By Story")
    ws.sheet_view.showGridLines = False

    by_us: dict[str, list[dict]] = {}
    for test in data["tests"]:
        us = _us_from_nodeid(test["nodeid"])
        by_us.setdefault(us, []).append(test)

    row = 1
    for us_key in sorted(by_us.keys()):
        tests = by_us[us_key]
        story_name = _US_MAP.get(us_key, us_key)
        passed = sum(1 for t in tests if t["outcome"] == "passed")
        total = len(tests)

        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = f"{us_key}  —  {story_name}  ({passed}/{total} passed)"
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = HEADER_FILL if passed == total else PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        # Column sub-headers
        _header_row(ws, row, ["Test Class", "Test Name", "Outcome", "Duration (s)", ""])
        row += 1

        for test in tests:
            parts = test["nodeid"].split("::")
            test_class = parts[1] if len(parts) >= 3 else "—"
            test_name = parts[-1]
            outcome = test["outcome"]
            duration = round(test.get("duration", 0.0), 4)

            ws.cell(row=row, column=1, value=test_class).alignment = Alignment(horizontal="left")
            ws.cell(row=row, column=2, value=test_name).alignment = Alignment(horizontal="left")
            fill = _status_fill(outcome)
            cell = ws.cell(row=row, column=3, value=outcome.upper())
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=duration).alignment = Alignment(horizontal="right")
            row += 1

        row += 1  # blank separator

    _auto_width(ws)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14


# ═══════════════════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════════════════


def main() -> None:
    src = Path("test_results.json")
    if not src.exists():
        raise FileNotFoundError(f"Run pytest first: {src} not found")

    data = json.loads(src.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _sheet_summary(wb, data)
    _sheet_all_tests(wb, data)
    _sheet_failures(wb, data)
    _sheet_by_story(wb, data)

    out = Path("test_results.xlsx")
    wb.save(out)
    total = data["summary"]["total"]
    passed = data["summary"]["passed"]
    failed = data["summary"].get("failed", 0)
    print(f"Saved -> {out.resolve()}")
    print(f"  {total} tests  |  {passed} passed  |  {failed} failed")
    print("  Sheets: Summary, All Tests, Failures, By Story")


if __name__ == "__main__":
    main()
