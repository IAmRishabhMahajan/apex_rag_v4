"""Run APEX-RAG benchmark evaluations and export results to eval_results.xlsx.

Datasets are downloaded once and cached to data/<benchmark>_examples.json so
future runs are instant (no internet needed after first run).

Usage:
    uv run python run_evals_to_excel.py                   # all 5 benchmarks
    uv run python run_evals_to_excel.py --benchmark nq    # one benchmark
    uv run python run_evals_to_excel.py --max-examples 100
    uv run python run_evals_to_excel.py --refresh         # re-download everything

Available benchmarks: hotpotqa  nq  triviaqa  ragbench  multihop
"""

from __future__ import annotations

import argparse
import json
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

# ── colour palette ─────────────────────────────────────────────────────────────
NAVY = PatternFill("solid", fgColor="1F3864")
BLUE = PatternFill("solid", fgColor="2F5496")
LBLUE = PatternFill("solid", fgColor="BDD7EE")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GREY = PatternFill("solid", fgColor="D9D9D9")
LGREY = PatternFill("solid", fgColor="F2F2F2")
WHITE = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
NORM = Font(size=10)


def _cell(ws, r, c, v, bold=False, fill=None, align="center", wrap=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def _header(ws, r, labels, fill=BLUE):
    for c, lbl in enumerate(labels, 1):
        cell = ws.cell(row=r, column=c, value=lbl)
        cell.font = HDR_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 22


def _auto_width(ws, mn=10, mx=50):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx, max(mn, w + 2))


def _score_fill(v: float) -> PatternFill:
    if v >= 0.6:
        return GREEN
    if v >= 0.3:
        return YELLOW
    return RED


# ── local dataset cache helpers ────────────────────────────────────────────────


def _cache_path(name: str, n: int) -> Path:
    return DATA_DIR / f"{name}_{n}.json"


def _load_cache(name: str, n: int) -> list[dict] | None:
    p = _cache_path(name, n)
    if p.exists():
        print(f"     Loading from local cache: {p.name}", flush=True)
        return json.loads(p.read_text(encoding="utf-8"))
    return None


def _save_cache(name: str, n: int, examples: list[dict]) -> None:
    p = _cache_path(name, n)
    p.write_text(json.dumps(examples, ensure_ascii=False, default=str), encoding="utf-8")
    print(f"     Saved {len(examples)} examples -> {p.name}", flush=True)


# ── HotpotQA ───────────────────────────────────────────────────────────────────


def _run_hotpotqa_on(examples: list[dict]) -> dict[str, Any]:
    from benchmarks._utils import (
        avg,
        best_exact_match,
        best_token_f1,
        make_item,
        retrieved_ids,
        safe_run_pipeline,
    )

    answer_ems, answer_f1s, sf_f1s, joint_f1s = [], [], [], []
    type_buckets: dict[str, dict[str, list]] = {}
    level_buckets: dict[str, dict[str, list]] = {}
    failures = 0
    t0 = time.perf_counter()

    for i, ex in enumerate(examples):
        query = ex["question"]
        gold_ans = ex["answer"]
        qtype = ex.get("type", "unknown")
        level = ex.get("level", "unknown")
        context = ex["context"]
        sf_titles = set(ex["supporting_facts"]["title"])

        items = []
        for title, sents in zip(context["title"], context["sentences"], strict=False):
            content = " ".join(sents).strip()
            if content:
                items.append(
                    make_item(
                        content=content,
                        source_id=title,
                        title=title,
                        url=f"https://en.wikipedia.org/wiki/{title.replace(' ', '_')}",
                        expert="search",
                        query=query,
                    )
                )

        result, _ = safe_run_pipeline(query, items, query_id=f"hpqa-{i}")
        if result is None:
            failures += 1
            continue

        pred = result.answer.text
        aem = best_exact_match(pred, [gold_ans])
        af1 = best_token_f1(pred, [gold_ans])
        ret_ids = retrieved_ids(result)

        predicted = set(ret_ids)
        tp = len(predicted & sf_titles)
        prec = tp / len(predicted) if predicted else 0.0
        rec = tp / len(sf_titles) if sf_titles else 1.0
        sff = 2 * prec * rec / (prec + rec) if (prec + rec) else 0.0
        jf1 = min(af1, sff)

        answer_ems.append(aem)
        answer_f1s.append(af1)
        sf_f1s.append(sff)
        joint_f1s.append(jf1)

        for bkt, key in [(type_buckets, qtype), (level_buckets, level)]:
            if key not in bkt:
                bkt[key] = {"answer_f1": [], "sf_f1": [], "joint_f1": []}
            bkt[key]["answer_f1"].append(af1)
            bkt[key]["sf_f1"].append(sff)
            bkt[key]["joint_f1"].append(jf1)

    return {
        "benchmark": "HotpotQA",
        "config": "distractor",
        "split": "validation",
        "evaluated": len(examples) - failures,
        "total": len(examples),
        "failures": failures,
        "elapsed_s": time.perf_counter() - t0,
        "answer_em": avg(answer_ems),
        "answer_f1": avg(answer_f1s),
        "sf_f1": avg(sf_f1s),
        "joint_f1": avg(joint_f1s),
        "by_type": {k: {m: avg(v) for m, v in vs.items()} for k, vs in type_buckets.items()},
        "by_level": {k: {m: avg(v) for m, v in vs.items()} for k, vs in level_buckets.items()},
    }


def run_hotpotqa(max_examples: int, refresh: bool) -> dict[str, Any]:
    CACHE = "hotpotqa"
    examples = None if refresh else _load_cache(CACHE, max_examples)

    if examples is None:
        print("     Downloading hotpotqa/hotpot_qa from HuggingFace...", flush=True)
        from datasets import load_dataset  # type: ignore[import-untyped]

        ds = load_dataset("hotpotqa/hotpot_qa", "distractor", split="validation")
        raw = list(ds.select(range(min(max_examples, len(ds)))))
        examples = [
            {
                "question": ex["question"],
                "answer": ex["answer"],
                "type": ex["type"],
                "level": ex["level"],
                "context": {
                    "title": list(ex["context"]["title"]),
                    "sentences": [list(s) for s in ex["context"]["sentences"]],
                },
                "supporting_facts": {
                    "title": list(ex["supporting_facts"]["title"]),
                    "sent_id": list(ex["supporting_facts"]["sent_id"]),
                },
            }
            for ex in raw
        ]
        _save_cache(CACHE, max_examples, examples)

    return _run_hotpotqa_on(examples)


# ── Natural Questions ──────────────────────────────────────────────────────────


def _run_nq_on(examples: list[dict]) -> dict[str, Any]:
    from benchmarks._utils import avg, best_exact_match, best_token_f1, make_item, safe_run_pipeline

    short_ems, short_f1s = [], []
    yn_correct = yn_total = no_answer_handled = no_answer_total = failures = 0
    type_buckets: dict[str, dict[str, list]] = {}
    t0 = time.perf_counter()

    for i, ex in enumerate(examples):
        query = ex["query"]
        passage = ex.get("passage", "")
        title = ex.get("title", f"nq-{i}")
        url = ex.get("url", "")
        short_answers = ex.get("short_answers", [])
        yn_answers = ex.get("yn_answers", [])

        has_answer = bool(short_answers or yn_answers)
        if not has_answer:
            no_answer_total += 1

        if not passage.strip():
            failures += 1
            continue

        items = [
            make_item(
                content=passage, source_id=title, title=title, url=url, expert="search", query=query
            )
        ]
        result, _ = safe_run_pipeline(query, items, query_id=f"nq-{i}")
        if result is None:
            failures += 1
            continue

        pred = result.answer.text
        q_type = "no_answer"

        if short_answers:
            q_type = "short"
            em = best_exact_match(pred, short_answers)
            f1 = best_token_f1(pred, short_answers)
            short_ems.append(em)
            short_f1s.append(f1)
        elif yn_answers:
            q_type = "yes_no"
            yn_total += 1
            p = pred.strip().lower()
            if any(p.startswith(yn) for yn in yn_answers):
                yn_correct += 1
        else:
            if result.answer.has_limitations:
                no_answer_handled += 1

        if q_type not in type_buckets:
            type_buckets[q_type] = {"em": [], "f1": []}
        if q_type == "short":
            type_buckets[q_type]["em"].append(em)  # type: ignore[possibly-undefined]
            type_buckets[q_type]["f1"].append(f1)  # type: ignore[possibly-undefined]


    return {
        "benchmark": "Natural Questions",
        "config": "standard",
        "split": "validation",
        "evaluated": len(examples) - failures,
        "total": len(examples),
        "failures": failures,
        "elapsed_s": time.perf_counter() - t0,
        "short_em": avg(short_ems),
        "short_f1": avg(short_f1s),
        "yes_no_accuracy": yn_correct / yn_total if yn_total else 0.0,
        "no_answer_handling": no_answer_handled / no_answer_total if no_answer_total else 0.0,
        "by_type": {k: {m: avg(v) for m, v in vs.items()} for k, vs in type_buckets.items()},
    }


def run_nq(max_examples: int, refresh: bool) -> dict[str, Any]:
    CACHE = "nq"
    examples = None if refresh else _load_cache(CACHE, max_examples)

    if examples is None:
        import itertools

        print(
            "     Streaming google-research-datasets/natural_questions (validation)...", flush=True
        )
        from datasets import load_dataset  # type: ignore[import-untyped]

        ds = load_dataset(
            "google-research-datasets/natural_questions",
            split="validation",
            streaming=True,
        )
        MAX_TOKENS = 300
        examples = []
        for ex in itertools.islice(ds, max_examples):
            doc = ex["document"]
            anns = ex.get("annotations", {})
            query = (
                ex["question"]["text"] if isinstance(ex["question"], dict) else str(ex["question"])
            )

            tokens_field = doc.get("tokens", {}) if isinstance(doc, dict) else {}
            token_list = tokens_field.get("token", [])
            is_html_list = tokens_field.get("is_html", [])
            text_tokens = [t for t, h in zip(token_list, is_html_list, strict=False) if not h]
            passage = " ".join(text_tokens[:MAX_TOKENS])

            short_answers: list[str] = []
            yn_answers: list[str] = []
            if isinstance(anns, dict):
                for sa_group in anns.get("short_answers", []):
                    if isinstance(sa_group, dict):
                        for text in sa_group.get("text", []) or []:
                            if text and str(text).strip():
                                short_answers.append(str(text))
                for yn in anns.get("yes_no_answer", []):
                    if yn in (1, "YES", "yes"):
                        yn_answers.append("yes")
                    elif yn in (0, "NO", "no"):
                        yn_answers.append("no")

            examples.append(
                {
                    "query": query,
                    "passage": passage,
                    "title": doc.get("title", "") if isinstance(doc, dict) else "",
                    "url": doc.get("url", "") if isinstance(doc, dict) else "",
                    "short_answers": short_answers,
                    "yn_answers": yn_answers,
                }
            )

        _save_cache(CACHE, max_examples, examples)

    return _run_nq_on(examples)


# ── TriviaQA ───────────────────────────────────────────────────────────────────


def _run_triviaqa_on(examples: list[dict]) -> dict[str, Any]:
    from benchmarks._utils import (
        avg,
        best_exact_match,
        best_token_f1,
        compute_recall_at_k,
        make_item,
        retrieved_ids,
        safe_run_pipeline,
    )

    ems, f1s, recalls = [], [], []
    failures = 0
    t0 = time.perf_counter()

    for i, ex in enumerate(examples):
        query = ex["question"]
        aliases = ex.get("aliases", [])
        items_data = ex.get("items", [])

        if not aliases:
            failures += 1
            continue

        items = [
            make_item(
                content=it["content"],
                source_id=it["source_id"],
                title=it["title"],
                url=it["url"],
                expert="search",
                query=query,
            )
            for it in items_data
            if it.get("content", "").strip()
        ]
        if not items:
            failures += 1
            continue

        result, _ = safe_run_pipeline(query, items, query_id=f"trivia-{i}")
        if result is None:
            failures += 1
            continue

        pred = result.answer.text
        ems.append(best_exact_match(pred, aliases))
        f1s.append(best_token_f1(pred, aliases))

        gold_ids = [it["source_id"] for it in items_data if it.get("source_id")]
        ret = retrieved_ids(result)
        recalls.append(compute_recall_at_k(ret, gold_ids, k=5))

    return {
        "benchmark": "TriviaQA",
        "config": "rc",
        "split": "validation",
        "evaluated": len(examples) - failures,
        "total": len(examples),
        "failures": failures,
        "elapsed_s": time.perf_counter() - t0,
        "em": avg(ems),
        "f1": avg(f1s),
        "retrieval_recall": avg(recalls),
    }


def run_triviaqa(max_examples: int, refresh: bool) -> dict[str, Any]:
    CACHE = "triviaqa"
    examples = None if refresh else _load_cache(CACHE, max_examples)

    if examples is None:
        import itertools

        print("     Streaming mandarjoshi/trivia_qa rc (validation)...", flush=True)
        from datasets import load_dataset  # type: ignore[import-untyped]

        ds = load_dataset("mandarjoshi/trivia_qa", "rc", split="validation", streaming=True)

        MAX_CHARS = 1500
        examples = []
        for ex in itertools.islice(ds, max_examples):
            answer = ex.get("answer", {})
            aliases: list[str] = []
            for field in ("value", "aliases", "normalized_aliases", "normalized_value"):
                val = answer.get(field)
                if isinstance(val, list):
                    aliases.extend(v for v in val if v)
                elif val:
                    aliases.append(str(val))
            aliases = list(dict.fromkeys(a for a in aliases if a))

            pages = ex.get("entity_pages") or ex.get("search_results") or {}
            contexts = pages.get("wiki_context") or pages.get("search_context") or []
            titles = pages.get("title") or [""] * len(contexts)
            urls = pages.get("url") or [""] * len(contexts)

            items_data = []
            for ctx, ttl, url in zip(contexts, titles, urls, strict=False):
                content = (ctx or "")[:MAX_CHARS].strip()
                if content:
                    items_data.append(
                        {
                            "content": content,
                            "source_id": url or ttl or f"src-{len(items_data)}",
                            "title": str(ttl),
                            "url": str(url),
                        }
                    )

            examples.append(
                {
                    "question": ex["question"],
                    "aliases": aliases,
                    "items": items_data,
                }
            )

        _save_cache(CACHE, max_examples, examples)

    return _run_triviaqa_on(examples)


# ── RAGBench ───────────────────────────────────────────────────────────────────


def _run_ragbench_subset_on(subset_name: str, examples: list[dict]) -> dict[str, Any]:
    from benchmarks._utils import avg, make_item, safe_run_pipeline

    adherences, support_rates, has_lim = [], [], []
    failures = 0

    for i, ex in enumerate(examples):
        query = ex.get("query", "")
        docs = ex.get("docs", [])

        if not query or not docs:
            failures += 1
            continue

        items = [
            make_item(
                content=d,
                source_id=f"ragbench-{subset_name}-{j}",
                title=f"Document {j + 1}",
                url="",
                expert="search",
                query=query,
            )
            for j, d in enumerate(docs)
            if d.strip()
        ]
        if not items:
            failures += 1
            continue

        result, _ = safe_run_pipeline(query, items, query_id=f"rb-{subset_name}-{i}")
        if result is None:
            failures += 1
            continue

        eval_result = result.eval_result
        claims = getattr(eval_result, "claims", None)
        sr = getattr(claims, "support_rate", 0.0) if claims else 0.0
        support_rates.append(sr)
        adherences.append(sr)
        has_lim.append(1.0 if result.answer.has_limitations else 0.0)

    return {
        "subset": subset_name,
        "evaluated": len(examples) - failures,
        "total": len(examples),
        "adherence_rate": avg(adherences),
        "claim_support_rate": avg(support_rates),
        "has_lim_rate": avg(has_lim),
        "failures": failures,
    }


def run_ragbench(
    max_examples: int, refresh: bool, subsets: list[str] | None = None
) -> dict[str, Any]:
    from benchmarks._utils import avg

    if subsets is None:
        subsets = ["covidqa", "hotpotqa", "pubmedqa", "finqa", "cuad"]

    all_examples: dict[str, list[dict]] = {}
    for subset in subsets:
        CACHE = f"ragbench_{subset}"
        cached = None if refresh else _load_cache(CACHE, max_examples)
        if cached is not None:
            all_examples[subset] = cached
        else:
            print(f"     Downloading galileo-ai/ragbench [{subset}]...", flush=True)
            try:
                from datasets import load_dataset  # type: ignore[import-untyped]

                ds = load_dataset("galileo-ai/ragbench", subset, split="train")
                raw = list(ds.select(range(min(max_examples, len(ds)))))
                exs = [
                    {
                        "query": r.get("question", "") or r.get("query", ""),
                        "docs": r.get("documents", [])
                        if isinstance(r.get("documents"), list)
                        else ([r["documents"]] if r.get("documents") else []),
                    }
                    for r in raw
                ]
                _save_cache(CACHE, max_examples, exs)
                all_examples[subset] = exs
            except Exception as exc:
                print(f"     [WARN] {subset}: {exc}", flush=True)

    by_subset: dict[str, dict] = {}
    total_examples = failures = 0
    all_adherences: list[float] = []
    all_supports: list[float] = []

    for subset, examples in all_examples.items():
        print(f"     Evaluating ragbench/{subset} ({len(examples)} examples)...", flush=True)
        sd = _run_ragbench_subset_on(subset, examples)
        by_subset[subset] = sd
        total_examples += sd["total"]
        failures += sd["failures"]
        if sd["evaluated"] > 0:
            all_adherences.append(sd["adherence_rate"])
            all_supports.append(sd["claim_support_rate"])

    return {
        "benchmark": "RAGBench",
        "config": f"{len(subsets)} subsets",
        "split": "train",
        "evaluated": total_examples - failures,
        "total": total_examples,
        "failures": failures,
        "elapsed_s": 0.0,
        "overall_adherence": avg(all_adherences),
        "overall_support_rate": avg(all_supports),
        "by_subset": by_subset,
    }


# ── MultiHop-RAG ───────────────────────────────────────────────────────────────


def _run_multihop_on(examples: list[dict]) -> dict[str, Any]:
    from benchmarks._utils import (
        avg,
        best_exact_match,
        best_token_f1,
        compute_mrr,
        compute_precision_at_k,
        compute_recall_at_k,
        make_item,
        retrieved_ids,
        safe_run_pipeline,
    )

    mrrs, recalls, precisions, ems, f1s = [], [], [], [], []
    null_correct = null_total = failures = 0
    type_buckets: dict[str, dict[str, list]] = {}
    K = 5
    t0 = time.perf_counter()

    for i, ex in enumerate(examples):
        query = ex["query"]
        gold_answer = ex.get("answer", "")
        q_type = ex.get("question_type", "unknown")
        is_null = q_type == "null"
        evidence = ex.get("evidence", [])

        items = [
            make_item(
                content=ev["content"],
                source_id=ev["source_id"],
                title=ev.get("title", ""),
                url=ev.get("url", ""),
                expert="search",
                query=query,
            )
            for ev in evidence
            if ev.get("content", "").strip()
        ]
        if not items:
            failures += 1
            continue

        result, _ = safe_run_pipeline(query, items, query_id=f"mhop-{i}")
        if result is None:
            failures += 1
            continue

        pred = result.answer.text
        ret_ids = retrieved_ids(result)
        gold_ids = [ev["source_id"] for ev in evidence]

        mrrs.append(compute_mrr(ret_ids, gold_ids, k=K))
        recalls.append(compute_recall_at_k(ret_ids, gold_ids, k=K))
        precisions.append(compute_precision_at_k(ret_ids, gold_ids, k=K))

        if is_null:
            null_total += 1
            if result.answer.has_limitations:
                null_correct += 1
        else:
            em = best_exact_match(pred, [gold_answer])
            f1 = best_token_f1(pred, [gold_answer])
            ems.append(em)
            f1s.append(f1)

        if q_type not in type_buckets:
            type_buckets[q_type] = {"answer_f1": [], "mrr": [], "recall": []}
        type_buckets[q_type]["mrr"].append(mrrs[-1])
        type_buckets[q_type]["recall"].append(recalls[-1])
        if not is_null:
            type_buckets[q_type]["answer_f1"].append(f1)  # type: ignore[possibly-undefined]

    return {
        "benchmark": "MultiHop-RAG",
        "config": f"k={K}",
        "split": "train",
        "evaluated": len(examples) - failures,
        "total": len(examples),
        "failures": failures,
        "elapsed_s": time.perf_counter() - t0,
        "mrr": avg(mrrs),
        "recall_at_k": avg(recalls),
        "precision_at_k": avg(precisions),
        "answer_em": avg(ems),
        "answer_f1": avg(f1s),
        "null_handling_rate": null_correct / null_total if null_total else 0.0,
        "by_type": {k2: {m: avg(v) for m, v in vs.items()} for k2, vs in type_buckets.items()},
    }


def run_multihop(max_examples: int, refresh: bool) -> dict[str, Any]:
    CACHE = "multihop_rag"
    examples = None if refresh else _load_cache(CACHE, max_examples)

    if examples is None:
        import itertools

        print("     Streaming yixuantt/MultiHopRAG (train)...", flush=True)
        from datasets import load_dataset  # type: ignore[import-untyped]

        ds = load_dataset("yixuantt/MultiHopRAG", "MultiHopRAG", split="train", streaming=True)

        examples = []
        for ex in itertools.islice(ds, max_examples):
            ev_list = ex.get("evidence_list", []) or []
            evidence = []
            for ev in ev_list:
                facts = ev.get("facts", [])
                content = " ".join(facts) if facts else ev.get("title", "")
                source_id = ev.get("source", ev.get("title", f"src-{len(evidence)}"))
                if content.strip():
                    evidence.append(
                        {
                            "content": content,
                            "source_id": str(source_id),
                            "title": ev.get("title", ""),
                            "url": str(source_id) if str(source_id).startswith("http") else "",
                        }
                    )
            examples.append(
                {
                    "query": ex.get("query", ""),
                    "answer": ex.get("answer", ""),
                    "question_type": ex.get("question_type", "unknown"),
                    "evidence": evidence,
                }
            )

        _save_cache(CACHE, max_examples, examples)

    return _run_multihop_on(examples)


# ── Excel sheet builders ───────────────────────────────────────────────────────


def _sheet_summary(wb, results: list[dict], run_ts: datetime, total_s: float):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "APEX-RAG v4  -  Benchmark Evaluation Results"
    c.font = Font(bold=True, size=16, color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Run: {run_ts.strftime('%Y-%m-%d  %H:%M:%S')}   |   Total elapsed: {total_s:.1f}s"
    c.font = Font(italic=True, size=10, color="595959")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    HEADERS = [
        "Benchmark",
        "Config",
        "Evaluated",
        "Failures",
        "Primary Metric",
        "Score",
        "Secondary Metric",
        "Score",
    ]
    _header(ws, 4, HEADERS)

    primary = {
        "HotpotQA": ("Joint F1", "joint_f1"),
        "Natural Questions": ("Short-Answer F1", "short_f1"),
        "TriviaQA": ("F1 (alias-matched)", "f1"),
        "RAGBench": ("Overall Adherence", "overall_adherence"),
        "MultiHop-RAG": ("Answer F1", "answer_f1"),
    }
    secondary = {
        "HotpotQA": ("Answer F1", "answer_f1"),
        "Natural Questions": ("Short-Answer EM", "short_em"),
        "TriviaQA": ("EM (alias-matched)", "em"),
        "RAGBench": ("Claim Support Rate", "overall_support_rate"),
        "MultiHop-RAG": ("MRR@5", "mrr"),
    }

    r = 5
    for d in results:
        bname = d["benchmark"]
        pm_l, pm_k = primary.get(bname, ("--", ""))
        sm_l, sm_k = secondary.get(bname, ("--", ""))
        pm_val = float(d.get(pm_k) or 0)
        sm_val = float(d.get(sm_k) or 0)

        _cell(ws, r, 1, bname, bold=True, align="left", fill=LGREY)
        _cell(ws, r, 2, d.get("config", ""), align="center", fill=LGREY)
        _cell(ws, r, 3, d["evaluated"], align="center", fill=LGREY)
        _cell(ws, r, 4, d["failures"], align="center", fill=RED if d["failures"] > 0 else GREEN)
        _cell(ws, r, 5, pm_l, align="left", fill=LGREY)
        c = _cell(ws, r, 6, round(pm_val, 4), align="center", fill=_score_fill(pm_val))
        c.number_format = "0.000"
        _cell(ws, r, 7, sm_l, align="left", fill=LGREY)
        c = _cell(ws, r, 8, round(sm_val, 4), align="center", fill=_score_fill(sm_val))
        c.number_format = "0.000"
        ws.row_dimensions[r].height = 20
        r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["G"].width = 24


def _sheet_hotpotqa(wb, d: dict):
    ws = wb.create_sheet("HotpotQA")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"HotpotQA [{d['config']}]  -  {d['split']} split"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    metrics = [
        ("Examples evaluated", d["evaluated"]),
        ("Failures", d["failures"]),
        ("Answer EM", d["answer_em"]),
        ("Answer F1", d["answer_f1"]),
        ("Supporting-Fact F1", d["sf_f1"]),
        ("Joint F1", d["joint_f1"]),
        ("Elapsed (s)", round(d["elapsed_s"], 2)),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (lbl, val) in enumerate(metrics, 4):
        _cell(ws, i, 1, lbl, bold=True, align="left", fill=LGREY)
        fill = _score_fill(val) if isinstance(val, float) and "Elapsed" not in lbl else None
        c = _cell(ws, i, 2, round(val, 4) if isinstance(val, float) else val, fill=fill)
        if isinstance(val, float) and "Elapsed" not in lbl:
            c.number_format = "0.000"
    r = 4 + len(metrics) + 2
    for bkt_name, bkt in [
        ("By Question Type", d.get("by_type", {})),
        ("By Difficulty Level", d.get("by_level", {})),
    ]:
        if not bkt:
            continue
        ws.cell(row=r, column=1, value=bkt_name).font = BOLD
        r += 1
        _header(ws, r, [bkt_name.split()[-1], "Answer F1", "SF F1", "Joint F1"])
        r += 1
        for key, scores in bkt.items():
            _cell(ws, r, 1, key, align="left")
            for ci, m in enumerate(["answer_f1", "sf_f1", "joint_f1"], 2):
                v = float(scores.get(m) or 0)
                c = _cell(ws, r, ci, round(v, 4), fill=_score_fill(v))
                c.number_format = "0.000"
            r += 1
        r += 1
    _auto_width(ws)


def _sheet_nq(wb, d: dict):
    ws = wb.create_sheet("Natural Questions")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"Natural Questions  -  {d['split']} split"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    metrics = [
        ("Examples evaluated", d["evaluated"]),
        ("Failures", d["failures"]),
        ("Short-Answer EM", d["short_em"]),
        ("Short-Answer F1", d["short_f1"]),
        ("Yes/No Accuracy", d["yes_no_accuracy"]),
        ("No-Answer Handling", d["no_answer_handling"]),
        ("Elapsed (s)", round(d["elapsed_s"], 2)),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (lbl, val) in enumerate(metrics, 4):
        _cell(ws, i, 1, lbl, bold=True, align="left", fill=LGREY)
        fill = _score_fill(val) if isinstance(val, float) and "Elapsed" not in lbl else None
        c = _cell(ws, i, 2, round(val, 4) if isinstance(val, float) else val, fill=fill)
        if isinstance(val, float) and "Elapsed" not in lbl:
            c.number_format = "0.000"
    r = 4 + len(metrics) + 2
    if d.get("by_type"):
        ws.cell(row=r, column=1, value="By Answer Type").font = BOLD
        r += 1
        type_keys = list(next(iter(d["by_type"].values())).keys())
        _header(ws, r, ["Answer Type"] + type_keys)
        r += 1
        for qtype, scores in d["by_type"].items():
            _cell(ws, r, 1, qtype, align="left")
            for ci, key in enumerate(type_keys, 2):
                v = float(scores.get(key) or 0)
                c = _cell(ws, r, ci, round(v, 4), fill=_score_fill(v))
                c.number_format = "0.000"
            r += 1
    _auto_width(ws)


def _sheet_triviaqa(wb, d: dict):
    ws = wb.create_sheet("TriviaQA")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"].value = f"TriviaQA [{d['config']}]  -  {d['split']} split"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    metrics = [
        ("Examples evaluated", d["evaluated"]),
        ("Failures", d["failures"]),
        ("EM (alias-matched)", d["em"]),
        ("F1 (alias-matched)", d["f1"]),
        ("Retrieval Recall@5", d["retrieval_recall"]),
        ("Elapsed (s)", round(d["elapsed_s"], 2)),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (lbl, val) in enumerate(metrics, 4):
        _cell(ws, i, 1, lbl, bold=True, align="left", fill=LGREY)
        fill = _score_fill(val) if isinstance(val, float) and "Elapsed" not in lbl else None
        c = _cell(ws, i, 2, round(val, 4) if isinstance(val, float) else val, fill=fill)
        if isinstance(val, float) and "Elapsed" not in lbl:
            c.number_format = "0.000"
    _auto_width(ws)


def _sheet_ragbench(wb, d: dict):
    ws = wb.create_sheet("RAGBench")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"RAGBench [{d['config']}]  -  {d['split']} split"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    metrics = [
        ("Examples evaluated", d["evaluated"]),
        ("Failures", d["failures"]),
        ("Overall Adherence", d["overall_adherence"]),
        ("Overall Support Rate", d["overall_support_rate"]),
        ("Elapsed (s)", round(d["elapsed_s"], 2)),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (lbl, val) in enumerate(metrics, 4):
        _cell(ws, i, 1, lbl, bold=True, align="left", fill=LGREY)
        fill = _score_fill(val) if isinstance(val, float) and "Elapsed" not in lbl else None
        c = _cell(ws, i, 2, round(val, 4) if isinstance(val, float) else val, fill=fill)
        if isinstance(val, float) and "Elapsed" not in lbl:
            c.number_format = "0.000"
    r = 4 + len(metrics) + 2
    if d.get("by_subset"):
        ws.cell(row=r, column=1, value="By Domain Subset").font = BOLD
        r += 1
        _header(
            ws,
            r,
            ["Subset", "Evaluated", "Failures", "Adherence", "Claim Support", "Has Limitations"],
        )
        r += 1
        for name, sd in d["by_subset"].items():
            _cell(ws, r, 1, name, align="left", bold=True)
            _cell(ws, r, 2, sd["evaluated"], align="center")
            _cell(
                ws, r, 3, sd["failures"], align="center", fill=RED if sd["failures"] > 0 else GREEN
            )
            for ci, key in enumerate(["adherence_rate", "claim_support_rate", "has_lim_rate"], 4):
                v = float(sd.get(key) or 0)
                c = _cell(ws, r, ci, round(v, 4), fill=_score_fill(v))
                c.number_format = "0.000"
            r += 1
    _auto_width(ws)


def _sheet_multihop(wb, d: dict):
    ws = wb.create_sheet("MultiHop-RAG")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"MultiHop-RAG [{d['config']}]  -  {d['split']} split"
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    metrics = [
        ("Examples evaluated", d["evaluated"]),
        ("Failures", d["failures"]),
        ("MRR@5", d["mrr"]),
        ("Recall@5", d["recall_at_k"]),
        ("Precision@5", d["precision_at_k"]),
        ("Answer EM", d["answer_em"]),
        ("Answer F1", d["answer_f1"]),
        ("Null Handling Rate", d["null_handling_rate"]),
        ("Elapsed (s)", round(d["elapsed_s"], 2)),
    ]
    _header(ws, 3, ["Metric", "Value"])
    for i, (lbl, val) in enumerate(metrics, 4):
        _cell(ws, i, 1, lbl, bold=True, align="left", fill=LGREY)
        fill = _score_fill(val) if isinstance(val, float) and "Elapsed" not in lbl else None
        c = _cell(ws, i, 2, round(val, 4) if isinstance(val, float) else val, fill=fill)
        if isinstance(val, float) and "Elapsed" not in lbl:
            c.number_format = "0.000"
    r = 4 + len(metrics) + 2
    if d.get("by_type"):
        ws.cell(row=r, column=1, value="By Question Type").font = BOLD
        r += 1
        type_keys = list(next(iter(d["by_type"].values())).keys())
        _header(ws, r, ["Question Type"] + type_keys)
        r += 1
        for qtype, scores in d["by_type"].items():
            _cell(ws, r, 1, qtype, align="left")
            for ci, key in enumerate(type_keys, 2):
                v = float(scores.get(key) or 0)
                c = _cell(ws, r, ci, round(v, 4), fill=_score_fill(v))
                c.number_format = "0.000"
            r += 1
    _auto_width(ws)


def _save_excel(results: list[dict], out: Path, run_ts: datetime, total_s: float):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, results, run_ts, total_s)
    for d in results:
        b = d["benchmark"]
        if b == "HotpotQA":
            _sheet_hotpotqa(wb, d)
        elif b == "Natural Questions":
            _sheet_nq(wb, d)
        elif b == "TriviaQA":
            _sheet_triviaqa(wb, d)
        elif b == "RAGBench":
            _sheet_ragbench(wb, d)
        elif b == "MultiHop-RAG":
            _sheet_multihop(wb, d)
    wb.save(out)


# ── main ───────────────────────────────────────────────────────────────────────

RUNNERS = {
    "hotpotqa": ("HotpotQA", run_hotpotqa),
    "nq": ("Natural Questions", run_nq),
    "triviaqa": ("TriviaQA", run_triviaqa),
    "ragbench": ("RAGBench", run_ragbench),
    "multihop": ("MultiHop-RAG", run_multihop),
}


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--benchmark",
        choices=[*RUNNERS.keys(), "all"],
        default="all",
        help="Which benchmark to run (default: all)",
    )
    parser.add_argument(
        "--max-examples",
        type=int,
        default=100,
        dest="max_examples",
        help="Examples per benchmark (default: 100)",
    )
    parser.add_argument(
        "--out", default="eval_results.xlsx", help="Output Excel file (default: eval_results.xlsx)"
    )
    parser.add_argument(
        "--refresh", action="store_true", help="Re-download datasets even if local cache exists"
    )
    parser.add_argument(
        "--subsets",
        nargs="+",
        help="RAGBench subsets (default: covidqa hotpotqa pubmedqa finqa cuad)",
    )
    args = parser.parse_args()

    MAX = args.max_examples
    targets = list(RUNNERS.keys()) if args.benchmark == "all" else [args.benchmark]
    run_ts = datetime.now()
    t0 = time.perf_counter()
    results: list[dict] = []
    out = Path(args.out)

    for idx, name in enumerate(targets, 1):
        label, fn = RUNNERS[name]
        print(f"\n[{idx}/{len(targets)}] {label} ({MAX} examples)...", flush=True)
        bt = time.perf_counter()
        try:
            kwargs: dict = {"max_examples": MAX, "refresh": args.refresh}
            if name == "ragbench" and args.subsets:
                kwargs["subsets"] = args.subsets
            d = fn(**kwargs)
            if name == "ragbench":
                d["elapsed_s"] = time.perf_counter() - bt
            results.append(d)
            print(
                f"   done  ({d['evaluated']} evaluated, {d['failures']} failures, "
                f"{d['elapsed_s']:.1f}s)",
                flush=True,
            )

            # Save Excel incrementally after each benchmark completes
            _save_excel(results, out, run_ts, time.perf_counter() - t0)
            print(
                f"   Excel updated -> {out.name}  ({len(results)}/{len(targets)} benchmarks)",
                flush=True,
            )

        except Exception as exc:
            print(f"   [ERROR] {label}: {exc}", flush=True)
            traceback.print_exc()

    total_s = time.perf_counter() - t0
    if results:
        _save_excel(results, out, run_ts, total_s)

    print(f"\n{'=' * 55}")
    print(f"  Benchmarks completed : {len(results)}/{len(targets)}")
    print(f"  Total elapsed        : {total_s:.1f}s")
    print(f"  Output file          : {out.resolve()}")
    print(f"  Dataset cache        : {DATA_DIR}/")
    print(f"{'=' * 55}")


if __name__ == "__main__":
    main()
